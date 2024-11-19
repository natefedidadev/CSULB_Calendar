from fastapi import FastAPI, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import uvicorn
from app.classes.month import CalMonth
from app.classes.year import CalYear, Calendar_Input
import base64
import io
from io import BytesIO
import time
import itertools
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from app.classes.year import DayType
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

app = FastAPI()

class Month(BaseModel):
    month: int = 1
    day_colors: list | None = ['red', 'blue']

class Base_Input(BaseModel):
    month: int
    day: int 
    year: int
    
@app.post("/calendar/test")
async def testThree(req: Month):
    return {
                "message": "hello world",
                "day_colors": req.day_colors
            }

@app.post("/calendar/table")
async def testTwo(req: Month):
    year = CalYear()
    awd_fall = [12, 21, 22, 19, 19]
    id_fall = [6, 21, 22, 16, 9]
    months = ['Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    img = year.create_months_table(awd=awd_fall, id=id_fall, months=months)
    
    img_bytes_io = io.BytesIO()
    img.save(img_bytes_io, format='PNG')
    
    img_bytes_io.seek(0)

    return StreamingResponse(img_bytes_io, media_type="image/png")

@app.post("/calendar/month")
async def testOne(req: Calendar_Input):
    cmonth = CalMonth(req.year, req.month)
    img = cmonth.draw(req.width)
    img_bytes_io = io.BytesIO()
    img.save(img_bytes_io, format='PNG')
    img_bytes_io.seek(0)
    
    return StreamingResponse(img_bytes_io, media_type="image/png")

@app.post("/calendar/build_year")
async def build_year(req: Calendar_Input):    
    start = time.perf_counter()
    calyear = CalYear(req)
    if calyear.valid:
        result_image = calyear.gen_schedule()
    
        img_bytes_io = io.BytesIO()
        result_image.save(img_bytes_io, format='PNG')
        
        img_bytes_io.seek(0)
        
        end = time.perf_counter()
        print("Elapsed (time) = {}s".format((end - start)))
        return StreamingResponse(img_bytes_io, media_type="image/png")
    return None

@app.post("/calendar/build_years_test")
def build_years_request2(req: Calendar_Input):    
    result = []
    calyear = CalYear(req)
    cnt = 0

    awd = 172
    id = 148
    convo_day = 3
    winter_sess = 13

    print(f"\nGenerating calendar {cnt}")
    start = time.perf_counter()
    result_image, a = calyear.gen_schedule(awd, id, convo_day, winter_sess)
    cnt += 1
    if result_image is None:
        pass
    else:
        img_bytes_io = io.BytesIO()
        result_image.save(img_bytes_io, format='PNG')
        img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
        returndict = {"image": img_str}
        result.append(returndict)
    end = time.perf_counter()
    print("Elapsed (time) = {}s".format((end - start)))
        
    return result

@app.post("/calendar/build_years_sequential")
def build_years_request2(req: Calendar_Input):    
    result = []
    calyear = CalYear(req)
    result_dict = {}
    cnt = 0
    for awd in range(170, 181):
        for id in range(145, 150):
            for convo_day in range(2, 6):
                for winter_sess in range(12, 16):
                    print(f"\nGenerating calendar {cnt}")
                    start = time.perf_counter()
                    result_image, md5_hash = calyear.gen_schedule(awd, id, convo_day, winter_sess)
                    if md5_hash in result_dict:
                        result_image = None
                        print("Duplicate calendar, discarded")
                    else:
                        result_dict[md5_hash] = True
                    cnt += 1
                    if result_image == None:
                        pass
                    else:
                        img_bytes_io = io.BytesIO()
                        result_image.save(img_bytes_io, format='PNG')
                        img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
                        returndict = {"image": img_str}
                        result.append(returndict)
                    end = time.perf_counter()
                    print("Elapsed (time) = {}s".format((end - start)))
        
    return result

@app.post("/calendar/build_years")
def build_years_request(req: Calendar_Input):
    result_dict = {}
    results = []
    cnt = 0

    # Generate all combinations of parameters
    for awd, id_days, convo_day, winter_sess in itertools.product(
        range(170, 181), range(145, 150), range(2, 6), range(12, 16)
    ):
        calyear = CalYear(req)
        print(f"Processing calendar {cnt}")
        result_image, md5_hash = calyear.gen_schedule(awd, id_days, convo_day, winter_sess)
        if md5_hash in result_dict:
            print("Duplicate calendar, discarded")
        else:
            result_dict[md5_hash] = {
                "awd": awd,
                "id": id_days,
                "convo_day": convo_day,
                "winter_sess": winter_sess
            }
            if result_image:
                img_bytes_io = io.BytesIO()
                result_image.save(img_bytes_io, format='PNG')
                img_str = base64.b64encode(img_bytes_io.getvalue()).decode("utf-8")
                results.append({
                    "image": img_str,
                    "parameters": {
                        "awd": awd,
                        "id": id_days,
                        "convo_day": convo_day,
                        "winter_sess": winter_sess
                    },
                    "hash": md5_hash
                })
        cnt += 1

    return results



def generate_colored_excel_calendar(self):
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Academic Calendar"

    # Define the color map for each DayType
    day_type_colors = {
        DayType.NONE: "FFFFFF",  # White
        DayType.AWD: "CCE5FF",   # Light blue
        DayType.ID: "EFEF95",    # Yellow
        DayType.CONVOCATION: "50E3C2",
        DayType.FINALS: "F8CBAD",
        DayType.NO_CLASS_CAMPUS_OPEN: "AE76C7",
        DayType.COMMENCEMENT: "BD10E0",
        DayType.SUMMER_SESSION: "C5E0B4",
        DayType.WINTER_SESSION: "CFCFCF",
        DayType.HOLIDAY: "FFC107",
        DayType.VOID: "BDBDBD"
    }

    # Set the starting position for the calendar layout
    start_row = 2
    start_col = 1

    # Iterate through each month and populate the calendar in Excel
    for month_index, cal_month in enumerate(self.months):
        row_offset = start_row + (month_index // 3) * 9
        col_offset = start_col + (month_index % 3) * 8

        # Write the month name
        ws.cell(row=row_offset, column=col_offset, value=cal_month.get_title()).font = Font(bold=True)
        # Day headers
        for i, day_name in enumerate(["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"], start=col_offset):
            ws.cell(row=row_offset + 1, column=i, value=day_name).alignment = Alignment(horizontal="center")

        # Populate days with colors
        for week_row, week in enumerate(cal_month.cal, start=row_offset + 2):
            for day_col, day in enumerate(week, start=col_offset):
                if day == 0:
                    continue
                cell = ws.cell(row=week_row, column=day_col, value=day)
                date_obj = date(cal_month.year, cal_month.month, day)
                if date_obj in self.cal_dict:
                    day_type = self.cal_dict[date_obj]
                    cell.fill = PatternFill(start_color=day_type_colors[day_type], end_color=day_type_colors[day_type], fill_type="solid")
                    if day_type == DayType.HOLIDAY or day_type == DayType.COMMENCEMENT:
                        cell.font = Font(bold=True)

    # Add Legend
    legend_row = row_offset + 10
    for index, (day_type, color) in enumerate(day_type_colors.items()):
        ws.cell(row=legend_row, column=start_col + index * 2, value=day_type.name)
        ws.cell(row=legend_row, column=start_col + index * 2 + 1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Add summary
    summary_row = legend_row + 2
    ws.cell(row=summary_row, column=start_col, value="Academic Work Days = 172").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=start_col, value="Instructional Days = 148").font = Font(bold=True)

    # Save workbook
    wb.save("academic_calendar.xlsx")


@app.post("/calendar/download_excel_colored")
async def download_excel_colored(input_data: dict):
    # Initialize CalYear with inputs
    calendar_input = Calendar_Input(**input_data)
    calendar = CalYear(calendar_input)

    # Extract optional parameters
    awd = input_data.get('awd')
    id_days = input_data.get('id')
    convo_day = input_data.get('convo_day')
    winter_sess = input_data.get('winter_sess')

    # Generate the calendar schedule with the provided parameters
    if awd and id_days and convo_day and winter_sess:
        calendar.gen_schedule(awd, id_days, convo_day, winter_sess)
    else:
        calendar.gen_schedule()

    # Generate Excel file
    excel_buffer = calendar.generate_colored_excel_calendar()

    # Return the file content for download
    headers = {"Content-Disposition": "attachment; filename=academic_calendar.xlsx"}
    return StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
